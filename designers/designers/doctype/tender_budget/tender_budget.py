from __future__ import annotations

from pathlib import Path
import time

import frappe
from frappe import _
from frappe.model.document import Document
from frappe.utils.file_manager import get_file_path

MAX_COMPARE_CHANGES = 500
ALLOWED_BUDGET_FILE_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}


class TenderBudget(Document):
    PARENT_STATUS_BY_BUDGET_STATUS = {
        "Draft": "Budget Drafting",
        "Prapare": "Budget Drafting",
        "In Progress": "Budget Drafting",
        "Review 1 level": "Budget Director Review",
        "Review 2 level": "Budget CEO Review",
        "Approved": "Budget Approved",
        "Rejected": "Rejected",
        "Cancelled": "Budget Drafting",
    }

    def _validate_budget_file(self):
        if not self.budget_file:
            return

        ext = Path((self.budget_file or "").split("?", 1)[0]).suffix.lower()
        if ext not in ALLOWED_BUDGET_FILE_EXTENSIONS:
            frappe.throw(_("Budget File must be an Excel file (.xlsx, .xlsm, .xls)"))

    def _archive_previous_versions(self):
        if not self.tender_request or not self.name:
            return

        # Retry helps under concurrent edits from UI/list views.
        for attempt in range(3):
            try:
                frappe.db.sql(
                    """
                    update `tabTender Budget`
                    set status = 'Archived'
                    where tender_request = %s
                      and name != %s
                      and status != 'Archived'
                    """,
                    (self.tender_request, self.name),
                )
                return
            except frappe.QueryDeadlockError:
                if attempt == 2:
                    raise
                time.sleep(0.2)

    def validate(self):
        self._validate_budget_file()

        if not self.tender_request:
            return

        latest_version = frappe.db.sql(
            """
            select coalesce(max(version), 0)
            from `tabTender Budget`
            where tender_request = %s
            """,
            (self.tender_request,),
        )[0][0]

        # Any non-latest version must always stay archived.
        if self.version and int(self.version) < int(latest_version):
            self.status = "Archived"

    def before_insert(self):
        if not self.tender_request:
            frappe.throw("Tender Request is required")

        # Keep track of previous latest record for deterministic archive.
        self.flags.previous_latest_budget = frappe.db.get_value(
            "Tender Budget",
            {"tender_request": self.tender_request},
            "name",
            order_by="version desc",
        )

        max_version = frappe.db.sql(
            """
            select coalesce(max(version), 0)
            from `tabTender Budget`
            where tender_request = %s
            """,
            (self.tender_request,),
        )[0][0]

        self.version = int(max_version) + 1

    def after_insert(self):
        # Keep only the newest version active.
        # Every previous version for the same Tender Request is archived.
        previous_latest = self.flags.get("previous_latest_budget")
        if previous_latest:
            frappe.db.set_value("Tender Budget", previous_latest, "status", "Archived", update_modified=False)
        self._archive_previous_versions()

    def on_update(self):
        if self.tender_request and self.version:
            # If this document is not the latest one anymore, force archive.
            latest_version = frappe.db.sql(
                """
                select coalesce(max(version), 0)
                from `tabTender Budget`
                where tender_request = %s
                """,
                (self.tender_request,),
            )[0][0]
            if int(self.version) < int(latest_version) and self.status != "Archived":
                self.db_set("status", "Archived", update_modified=False)
            elif int(self.version) == int(latest_version):
                # Re-apply invariant on any update of the latest version.
                self._archive_previous_versions()

        if self.tender_request:
            frappe.db.set_value("Tender Request", self.tender_request, "budget_version", self.version, update_modified=False)
            parent_status = self.PARENT_STATUS_BY_BUDGET_STATUS.get(self.status)
            if parent_status:
                frappe.db.set_value("Tender Request", self.tender_request, "status", parent_status, update_modified=False)


def _get_budget_by_version(tender_request: str, version: int) -> dict:
    budget = frappe.db.get_value(
        "Tender Budget",
        {"tender_request": tender_request, "version": int(version)},
        ["name", "version", "budget_file", "status"],
        as_dict=True,
    )
    if not budget:
        frappe.throw(_("Tender Budget version {0} was not found").format(version))
    if not budget.budget_file:
        frappe.throw(_("Tender Budget version {0} has no attached Budget File").format(version))
    return budget


def _load_workbook(file_url: str):
    try:
        from openpyxl import load_workbook
    except ImportError:
        frappe.throw(_("openpyxl is not installed in the current environment"))

    ext = Path((file_url or "").split("?", 1)[0]).suffix.lower()
    if ext not in {".xlsx", ".xlsm"}:
        frappe.throw(_("Comparison supports only .xlsx/.xlsm files"))

    file_path = get_file_path(file_url)
    return load_workbook(filename=file_path, data_only=True, read_only=True)


def _compare_workbooks(old_wb, new_wb) -> dict:
    old_sheets = set(old_wb.sheetnames)
    new_sheets = set(new_wb.sheetnames)

    added_sheets = sorted(new_sheets - old_sheets)
    removed_sheets = sorted(old_sheets - new_sheets)

    changes = []

    for sheet in sorted(old_sheets & new_sheets):
        old_ws = old_wb[sheet]
        new_ws = new_wb[sheet]

        max_row = max(old_ws.max_row or 0, new_ws.max_row or 0)
        max_col = max(old_ws.max_column or 0, new_ws.max_column or 0)

        if max_row == 0 or max_col == 0:
            continue

        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                old_val = old_ws.cell(row=r, column=c).value
                new_val = new_ws.cell(row=r, column=c).value
                if old_val != new_val:
                    changes.append(
                        {
                            "sheet": sheet,
                            "cell": f"{new_ws.cell(row=r, column=c).coordinate}",
                            "old": "" if old_val is None else str(old_val),
                            "new": "" if new_val is None else str(new_val),
                        }
                    )
                    if len(changes) >= MAX_COMPARE_CHANGES:
                        return {
                            "added_sheets": added_sheets,
                            "removed_sheets": removed_sheets,
                            "changes": changes,
                            "truncated": True,
                        }

    return {
        "added_sheets": added_sheets,
        "removed_sheets": removed_sheets,
        "changes": changes,
        "truncated": False,
    }


@frappe.whitelist()
def compare_budget_versions(
    tender_request: str | None = None,
    from_version: int | None = None,
    to_version: int | None = None,
):
    if not tender_request:
        frappe.throw(_("Tender Request is required"))
    if not from_version or not to_version:
        frappe.throw(_("Both versions are required"))
    if int(from_version) == int(to_version):
        frappe.throw(_("Choose two different versions"))

    old_budget = _get_budget_by_version(tender_request, int(from_version))
    new_budget = _get_budget_by_version(tender_request, int(to_version))

    old_wb = _load_workbook(old_budget.budget_file)
    new_wb = _load_workbook(new_budget.budget_file)
    result = _compare_workbooks(old_wb, new_wb)

    return {
        "tender_request": tender_request,
        "from": old_budget,
        "to": new_budget,
        "summary": {
            "added_sheets": result["added_sheets"],
            "removed_sheets": result["removed_sheets"],
            "changes_count": len(result["changes"]),
            "truncated": result["truncated"],
            "max_changes": MAX_COMPARE_CHANGES,
        },
        "changes": result["changes"],
    }
